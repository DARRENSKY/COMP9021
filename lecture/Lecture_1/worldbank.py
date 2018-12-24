# Written by Eric Martin for COMP9021


'''
Extracts from the list of countries accessible at http://www.worldbank.org/en/country
the value of IBRD/IDA operations, when available, and creates a spreadsheet with those data.

'''

import os.path
import urllib.request
import sys

import bs4
import openpyxl


def convert_to_number(amount):
    units = {'thousand': 10 ** 3, 'million': 10 ** 6, 'billion': 10 ** 9, 'trillion': 10 ** 12}
    # amount is of the form "$v " followed by one of the units above,
    # possibly preceded or followed by spaces, with v a floating point value.
    for unit in units:
        if unit in amount:
            return int(float(amount.strip().lstrip('$').rstrip(unit)) * units[unit])

def countries_and_data():
    # Example of html code being matched:
    #     <li class="name-country">
    #        <span><a href="http://www.worldbank.org/en/country/afghanistan">Afghanistan
    #              </a>
    #        </span>
    #     </li>
    for country in top_page.select('.name-country > span > a'):
        country_name = country.getText()
        # See comment further down.
        if country_name == 'Serbia':
            continue
        try:
            with urllib.request.urlopen(country.get('href')) as overview_url:
                overview_page = bs4.BeautifulSoup(overview_url, 'html.parser')
                try:
                    # Example of html code being matched:
                    # <div class="_loop_card_footer">
                    #    <a href="http://www.worldbank.org/en/country/afghanistan/overview"
                    #                                               class="_loop_card_link">Overview
                    #    </a>
                    # </div>
                    data = overview_page.select('._loop_card_footer > a')[1]
                except IndexError:
                    print(f'Could not access the resource for {country_name}.')
                    continue
                try:
                    with urllib.request.urlopen(data.get('href')) as data_url:
                        data_page = bs4.BeautifulSoup(data_url, 'html.parser')
                        # Example of html code being matched:
                        # <div class="chart-summry" data-reactid="308">
                        #    <div data-reactid="309">
                        #       <em data-reactid="310">$250.00 million</em>
                        #       ...
                        #    </div>
                        # </div>
                        extracted_data = data_page.select('.chart-summry > div > em')
                        if not extracted_data:
                            print(f'Could not access the resource for {country_name}.')
                            continue
                        # We assume that the first member of extracted_data is for
                        # the indicator of interest.
                        # It is wrong for Serbia (for which there is only one,
                        # different indicator). It might be wrong for others too.
                        yield (country_name, convert_to_number(extracted_data[0].getText()))
                except urllib.error.HTTPError:
                    print(f'Could not access the resource for {country_name}.')
        except urllib.error.HTTPError:
            print(f'Could not access the resource for {country_name}.')


file_name = 'IBRD_IDA_operations.xlsx'
if os.path.isfile(file_name):
    print(f'A file named {file_name} already exists.') 
    print('You have to remove it first.')
    sys.exit()
try:
    with urllib.request.urlopen('http://www.worldbank.org/en/country') as top_url:
        top_page = bs4.BeautifulSoup(top_url, 'html.parser')
        workbook = openpyxl.Workbook()
        spreadsheet = workbook.active
        spreadsheet.title = 'World countries'
        spreadsheet['A1'] = 'Country'
        spreadsheet['B1'] = 'IBRD/IDA operations'
        for counter, (country, amount) in enumerate(countries_and_data(), 2):
            spreadsheet.cell(row = counter, column = 1).value = country
            spreadsheet.cell(row = counter, column = 2).value = amount
        workbook.save(file_name)
except urllib.error.HTTPError:
    print('Could not access the top resource.')
    sys.exit()
